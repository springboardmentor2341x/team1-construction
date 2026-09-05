import io
from datetime import datetime, timezone
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ReportExcelGenerator:

    @staticmethod
    def generate_excel(sheet_title: str, report_title: str, project_info: Dict[str, Any], summary_kpis: List[Dict[str, str]], table_headers: List[str], table_rows: List[List[str]], extra_notes: str = None) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title[:30]
        ws.views.sheetView[0].showGridLines = True

        # Typography & Styles
        title_font = Font(name="Calibri", size=15, bold=True, color="0F172A")
        subtitle_font = Font(name="Calibri", size=10, bold=True, color="D97706")
        header_font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        
        meta_label_font = Font(name="Calibri", size=9.5, bold=True, color="475569")
        meta_val_font = Font(name="Calibri", size=9.5, color="0F172A")
        
        kpi_label_font = Font(name="Calibri", size=9, bold=True, color="475569")
        kpi_val_font = Font(name="Calibri", size=13, bold=True, color="D97706")
        kpi_fill = PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # 1. Header Title Block
        ws.cell(row=1, column=1, value="BUILDTRACK CONSTRUCTION MANAGEMENT SYSTEM").font = subtitle_font
        ws.cell(row=2, column=1, value=report_title).font = title_font
        ws.cell(row=3, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = meta_val_font
        
        ws.append([])

        # 2. Project Metadata Block
        current_row = 5
        ws.cell(row=current_row, column=1, value="Project Code:").font = meta_label_font
        ws.cell(row=current_row, column=2, value=project_info.get("code", "N/A")).font = meta_val_font
        ws.cell(row=current_row, column=3, value="Project Name:").font = meta_label_font
        ws.cell(row=current_row, column=4, value=project_info.get("name", "N/A")).font = meta_val_font

        current_row += 1
        ws.cell(row=current_row, column=1, value="Category / Status:").font = meta_label_font
        ws.cell(row=current_row, column=2, value=f"{project_info.get('category', 'General')} ({project_info.get('status', 'Active')})").font = meta_val_font
        ws.cell(row=current_row, column=3, value="Project Manager:").font = meta_label_font
        ws.cell(row=current_row, column=4, value=project_info.get("pm_name", "N/A")).font = meta_val_font

        current_row += 2

        # 3. KPI Summary Block
        if summary_kpis:
            ws.cell(row=current_row, column=1, value="Executive KPI Summary").font = title_font
            current_row += 1
            col_idx = 1
            for kpi in summary_kpis:
                c1 = ws.cell(row=current_row, column=col_idx, value=kpi.get("label", ""))
                c1.font = kpi_label_font
                c1.fill = kpi_fill
                c1.alignment = Alignment(horizontal="center", vertical="center")
                
                c2 = ws.cell(row=current_row + 1, column=col_idx, value=kpi.get("value", ""))
                c2.font = kpi_val_font
                c2.fill = kpi_fill
                c2.alignment = Alignment(horizontal="center", vertical="center")
                
                col_idx += 1
            current_row += 3

        # 4. Extra Notice if present
        if extra_notes:
            c_note = ws.cell(row=current_row, column=1, value=f"Notice: {extra_notes}")
            c_note.font = Font(name="Calibri", size=9.5, italic=True, color="B45309")
            current_row += 2

        # 5. Data Table Headers
        if table_headers:
            ws.cell(row=current_row, column=1, value="Detailed Data Records").font = Font(name="Calibri", size=11, bold=True, color="0F172A")
            current_row += 1
            
            for col_idx, h_text in enumerate(table_headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=h_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            current_row += 1

            # Data Table Rows
            zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            for r_idx, row in enumerate(table_rows):
                for col_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = Font(name="Calibri", size=9.5, color="0F172A")
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    if r_idx % 2 == 1:
                        cell.fill = zebra_fill
                current_row += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        return excel_bytes
